import csv
import re
import os
from statistics import mean
from openpyxl import Workbook
from openpyxl.styles import Side, Border, Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.colors as colors
import pdfkit
from jinja2 import Environment, FileSystemLoader

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = DataSet.csv_universal_parser(file_name)

    @staticmethod
    def delete_tags(string):
        text = '; '.join(re.sub(r"<[^>]+>", '', string).split('\n'))
        return ' '.join(text.replace('\r', '').split())

    @staticmethod
    def csv_universal_parser(file_name):
        with open(file_name, encoding="utf_8_sig") as file:
            reader = csv.reader(file)

            try:
                list_naming = next(reader)
            except StopIteration:
                print("Пустой файл")
                quit()

            reader = [row for row in reader if len(row) == len(list_naming) and '' not in row]

            if len(reader) == 0:
                print("Нет данных")
                quit()

            vacancies = []

            for row in reader:
                rows_dict = {}
                for head, cell in zip(list_naming, row):
                    rows_dict[head] = DataSet.delete_tags(cell)
                vacancies.append(Vacancy([rows_dict['name'], Salary(
                    [rows_dict['salary_from'], rows_dict['salary_to'], rows_dict['salary_currency']]),
                                          rows_dict['area_name'], rows_dict['published_at']]))

            return vacancies


class Vacancy:
    def __init__(self, vacancy_list):
        self.name = vacancy_list[0]
        self.salary = vacancy_list[1]
        self.area_name = vacancy_list[2]
        self.published_at = int(vacancy_list[3][:4])

    def get_output_vacancy(self):
        return [self.name, self.salary.get_output_salary(), self.area_name, self.published_at]


class Salary:
    def __init__(self, salary_list):
        self.salary_from = float(salary_list[0])
        self.salary_to = float(salary_list[1])
        self.salary_currency = salary_list[2]

    def get_avg_salary_in_rub(self):
        return (self.salary_from + self.salary_to) * currency_to_rub[self.salary_currency] / 2


class InputConnect:
    @staticmethod
    def print_vacancies():
        file_name = input('Введите название файла: ')
        vacancy = input('Введите название профессии: ')
        vacancies = DataSet(file_name).vacancies_objects

        statistic = InputConnect.get_statistic(vacancies, vacancy)

        print(f"Динамика уровня зарплат по годам: {statistic[0]}")
        print(f"Динамика количества вакансий по годам: {statistic[1]}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {statistic[2]}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {statistic[3]}")
        print(f"Уровень зарплат по городам (в порядке убывания): {statistic[4]}")
        print(f"Доля вакансий по городам (в порядке убывания): {statistic[5]}")

        Report(vacancy).generate_xlsx(statistic)
        Graph(vacancy).draw_graphs(statistic)

    @staticmethod
    def get_statistic(vacancies, vacancy):
        total = len(vacancies)
        years, vacancy_years, cities = InputConnect.get_data_dicts(vacancies, vacancy)

        salary_per_years, count_per_years = InputConnect.get_output_dicts(years.items())
        salary_per_vacancy_years, count_per_vacancy_years = InputConnect.get_output_dicts(vacancy_years.items())
        statistic = [salary_per_years, count_per_years, salary_per_vacancy_years, count_per_vacancy_years]

        salary_per_cities, count_per_cities = InputConnect.get_output_dicts(cities.items(), int(total / 100))
        salary_per_cities = sorted(salary_per_cities.items(), key=lambda city: city[1], reverse=True)
        statistic.append({city[0]: city[1] for city in salary_per_cities[:min(len(salary_per_cities), 10)]})
        count_per_cities = sorted(count_per_cities.items(), key=lambda city: city[1], reverse=True)
        statistic.append(
            {city[0]: round(city[1] / total, 4) for city in count_per_cities[:min(len(count_per_cities), 10)]})

        return statistic

    @staticmethod
    def append_dict(dictionary, key, salary):
        if key in dictionary:
            dictionary[key].append(salary)
        else:
            dictionary[key] = [salary]

    @staticmethod
    def get_output_dicts(items, minimum=0):
        salary_dict = {}
        count_dict = {}

        for key, value in items:
            count = len(value)
            if count >= minimum:
                salary_dict[key] = int(mean(value)) if count > 0 else 0
                count_dict[key] = count

        return salary_dict, count_dict

    @staticmethod
    def get_data_dicts(vacancies, vac_name):
        list_years = set(vacancy.published_at for vacancy in vacancies)
        list_years = list(range(min(list_years), max(list_years) + 1))
        years = {year: [] for year in list_years}
        vac_name_years = {year: [] for year in list_years}
        cities = {}

        for vacancy in vacancies:
            salary = vacancy.salary.get_avg_salary_in_rub()
            year = vacancy.published_at
            InputConnect.append_dict(years, year, salary)
            InputConnect.append_dict(cities, vacancy.area_name, salary)
            if vac_name in vacancy.name:
                InputConnect.append_dict(vac_name_years, year, salary)

        return years, vac_name_years, cities


class Report:
    def __init__(self, vacancy):
        self.ws1_attributes = {
            'A1': 'Год',
            'B1': 'Средняя зарплата',
            'C1': f"Средняя зарплата - {vacancy}",
            'D1': 'Количество вакансий',
            'E1': f"Количество вакансий - {vacancy}",
        }

        self.ws2_attributes = {
            'A1': 'Город',
            'B1': 'Уровень зарплат',
            'D1': 'Город',
            'E1': 'Доля вакансий'
        }

    def generate_xlsx(self, statistic):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Статистика по годам'
        ws2 = wb.create_sheet('Статистика по городам')

        Report.create_ws1(ws1, self.ws1_attributes, statistic[0], statistic[1], statistic[2], statistic[3])
        Report.create_ws2(ws2, self.ws2_attributes, statistic[4], statistic[5])

        wb.save('report.xlsx')

    @staticmethod
    def create_ws1(ws, title, salary_per_years, count_per_years, salary_per_vacancy_years, count_per_vacancy_years):
        Report.create_title(ws, title)

        for key in salary_per_years.keys():
            ws.append([key,
                       salary_per_years[key], salary_per_vacancy_years[key],
                       count_per_years[key], count_per_vacancy_years[key]])

        Report.create_border(ws, f"A1:E{len(salary_per_years) + 1}")
        Report.correct_length(ws)

    @staticmethod
    def create_ws2(ws, title, salary_per_cities, count_per_cities):
        Report.create_title(ws, title)
        count = len(salary_per_cities)

        salaries = [[key, value] for key, value in salary_per_cities.items()]
        fractions = [[key, value] for key, value in count_per_cities.items()]

        for i in range(count):
            row = salaries[i] + [''] + fractions[i]
            ws.append(row)

        for i in range(2, count + 2):
            ws[f"E{i}"].number_format = FORMAT_PERCENTAGE_00

        Report.create_border(ws, f"A1:B{count + 1}")
        Report.create_border(ws, f"D1:E{count + 1}")
        Report.correct_length(ws)

    @staticmethod
    def create_title(ws, title):
        font = Font(bold=True)

        for key, value in title.items():
            ws[key] = value
            ws[key].font = font

    @staticmethod
    def create_border(ws, table_range):
        side = Side(border_style="thin", color="000000")
        border = Border(side, side, side, side)

        for row in ws[table_range]:
            for cell in row:
                cell.border = border

    @staticmethod
    def correct_length(ws):
        for i, column in enumerate(ws.iter_cols()):
            length = 0
            for cell in column:
                value = cell.value if cell.value is not None else ''
                length = max(length, len(str(value)))
            ws.column_dimensions[get_column_letter(i + 1)].width = length + 3 if length != 0 else 0


class Graph:
    def __init__(self, vacancy):
        self.vacancy = vacancy

    def draw_graphs(self, statistic):
        figure, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        Graph.create_diagram(ax1, statistic[0], statistic[1],
                             ['Средняя з/п', f'З/п {self.vacancy}'], 'Уровень зарплат по годам')
        Graph.create_diagram(ax2, statistic[2], statistic[3],
                             ['Количество вакансий', f"Количество вакансий {self.vacancy}"],
                             'Количество вакансий по годам')

        Graph.create_barGraph(ax3, statistic[4], 'Уровень зарплат по городам')
        Graph.create_pie(ax4, statistic[5], 'Доля вакансий по городам')

        figure.tight_layout()
        figure.set_size_inches(8, 6)
        figure.set_dpi(350)
        figure.savefig('graph.png', dpi=350)

        plt.show()

    @staticmethod
    def create_diagram(ax, total, per_vacancy, legends, title):
        width = 0.35
        captions = total.keys()
        points = range(len(captions))
        total_arguments = list(total.values())
        per_vacancy_arguments = list(per_vacancy.values())

        ax.set_title(title)
        ax.bar(list(map(lambda x: x - width / 2, points)), total_arguments, width, label=legends[0])
        ax.bar(list(map(lambda x: x + width / 2, points)), per_vacancy_arguments, width, label=legends[1])
        ax.legend(prop={'size': 8})
        ax.grid(axis='y')

        for label in ax.get_yticklabels():
            label.set_fontsize(8)

        ax.set_xticks(points, captions, fontsize=8, rotation=90)

    @staticmethod
    def create_barGraph(ax, salary_per_cities, title):
        cities = list(salary_per_cities.keys())
        y_pos = list(range(len(cities)))

        ax.set_title(title)
        ax.barh(y_pos, list(salary_per_cities.values()), align='center')
        ax.invert_yaxis()
        ax.grid(axis='x')

        for label in ax.get_xticklabels():
            label.set_fontsize(8)

        ax.set_yticks(y_pos, labels=cities, fontsize=6)

    @staticmethod
    def create_pie(ax, count_per_cities, title):
        cities = list(count_per_cities.keys()) + ['Другие']
        values = list(count_per_cities.values())

        ax.pie(values + [1 - sum(values)], labels=cities, textprops={'size': 6}, colors=colors.BASE_COLORS)
        ax.set_title(title)


class ReportPdf:
    def __init__(self, vacancy):
        self.vacancy = vacancy

    def create_pdf(self, statistic):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("template.html")
        config = pdfkit.configuration(wkhtmltopdf='/usr/local/bin/wkhtmltopdf')
        path = os.path.abspath('')
        stat_years = ReportPdf.create_stat_years(statistic[0], statistic[1], statistic[2], statistic[3])
        stat_cities_salary = [{'city': key, 'salary': value} for key, value in statistic[4]]
        stat_cities_fraction = [{'city': key, 'fraction': str(value * 100) + '%'} for key, value in statistic[5]]

        pdf_template = template.render({'path': path,
                                        'vacancy': self.vacancy,
                                        'rows1': stat_years,
                                        'rows2': stat_cities_salary,
                                        'rows3': stat_cities_fraction,
                                        'image_file': 'graph.png'})
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config)

    @staticmethod
    def create_stat_years(salary_per_years, count_per_years, salary_per_vacancy_years, count_per_vacancy_years):
        return [{'year': key,
                 'avg_salary': salary_per_years[key],
                 'avg_salary_vacancy': salary_per_vacancy_years[key],
                 'count': count_per_years[key],
                 'count_vacancy': count_per_vacancy_years[key]} for key in salary_per_years.keys()]


def main():
    InputConnect().print_vacancies()


if __name__ == '__main__':
    main()

# "/Users/mihail/Downloads/vacancies_by_year.csv"
